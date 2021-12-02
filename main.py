import os
import openpyxl as op
import numpy as np
import matplotlib.pyplot as plt
import time
import pandas as pd


# выбор всех уникальных строк из файла по указанному столбцу
def unique_label(sheet, column_char):
    count: int = 2
    j = column_char + str(count)
    list_of_unique_names = []

    while str(sheet[j].value) != "None":
        if sheet[j].value in list_of_unique_names:
            count += 1
            j = column_char + str(count)
            continue
        else:
            list_of_unique_names.append(sheet[j].value)
            count += 1
            j = column_char + str(count)

    return list_of_unique_names, count


# вывод инфы в консоль и добавление её в массив (счётчик, среднее, 2 перцентиля)
def print_info(count, sum, arr_of_values, errors):
    text = []
    text.append('Count: ' + str(count))
    print(text[0])
    text.append('Average: ' + str(round(sum / count, 2)))
    print(text[1])
    text.append('Percentile 95: ' + str(round(np.percentile(arr_of_values, 95), 2)))
    print(text[2])
    text.append('Percentile 99: ' + str(round(np.percentile(arr_of_values, 99), 2)))
    print(text[3])
    text.append('Max: ' + str(max(arr_of_values)))
    print(text[4])
    text.append('Min: ' + str(min(arr_of_values)))
    print(text[5])
    text.append('Errors: ' + str(errors/count*100) + '%')
    print(text[6])
    return text


# удаление слеша
def del_slashes(checked_string):
    while checked_string.find('/') != -1:
        checked_string = checked_string.replace('/', '-')
    return checked_string

#
def del_question(checked_string):
    while checked_string.find('?') != -1:
        checked_string = checked_string.replace('?', '')
    return checked_string

def del_colon(checked_string):
    while checked_string.find(':') != -1:
        checked_string = checked_string.replace(':', '')
    return checked_string

# временный файл для сортировки
temp_file = 'file.xlsx'

# блок ввода имени файла + проверка
name_of_file = input('Введите путь к файлу и имя файла: ')
print('Проверка файла...')

try:
    source_file = op.load_workbook(name_of_file)
except op.utils.exceptions.InvalidFileException:
    print('Не верное имя или формат файла. Поддерживаются файлы с расширениями .xlsx,.xlsm,.xltx,.xltm')
    exit()
except FileNotFoundError:
    print('Файл с указанным именем не найден')
    exit()
else:
    print('OK')

# блок ввода имени листа + проверка
name_of_sheet = input('Введите название листа: ')

try:
    sheet = source_file[name_of_sheet]
except KeyError:
    print('Листа с данным названием не существует')
    exit()
else:
    print('OK')

# блок ввода букв столбцов
col_elapse = input('Введите букву столбца с временем отклика: ')
col_t_s = input('Введите букву столбца с временными метками: ')
col_label = input('Введите букву столбца с лейблами: ')
col_errors = input('Введите букву столбца с ошибками: ')
start_label = col_label + str(2)
name_of_col_sort = sheet[col_t_s + str(1)].value

print(name_of_col_sort)

# задание и проверка частоты
frequency = int(input('Введите частоту дат на графиках (в пределах от 1 до 20): '))

while frequency > 20 or frequency < 1:
    print('Неверное значение частоты, попробуйте снова.')
    frequency = int(input('Введите частоту дат на графиках (в пределах от 1 до 20): '))

date = input('Введите дату теста (для названия): ')

# сортировка файла по таймстемпам
source_file = pd.ExcelFile(name_of_file)
sheet = source_file.parse(name_of_sheet)

# проверка правильности при вводе буквы столбца временных меток
# при удаче запись отсортированных данных в файл-копию
try:
    sheet = sheet.sort_values(by=name_of_col_sort)
    sheet.to_excel(temp_file, sheet_name=name_of_sheet, index=False)
except KeyError:
    print('Ошибка данных в столбце временных меток')
    exit()

#
result_file = op.Workbook()
result_file.create_sheet(title='result', index=0)
result_sheet = result_file['result']
result_sheet['A1'] = 'Роль'
result_sheet['B1'] = 'Наименование операции'
result_sheet['C1'] = 'Лейбл'
result_sheet['D1'] = 'Счётчик'
result_sheet['E1'] = 'Среднее'
result_sheet['F1'] = 'Максимум'
result_sheet['G1'] = 'Минимум'
result_sheet['H1'] = '95 перцентиль'
result_sheet['I1'] = '99 перцентиль'
result_sheet['J1'] = 'Ошибки'

# окончание работы с исходным файлом
source_file.close()

#
names_file = op.load_workbook('Names.xlsx')
names_sheet = names_file.active

# начало работы с копией данных
my_file = op.load_workbook(temp_file)
sheet = my_file[name_of_sheet]

# получаем счётчик и массив с уникальными лейблами
main_array, counter_of_lines = unique_label(sheet, col_label)

counter_for_writing = 2

# прогон кода для каждого лейбла
all_names = map(str, main_array)
for name in all_names:

    pass_counter = 0
    summa = 0
    sum_of_errors = 0
    arr_of_elapse = []
    arr_of_ts = []

    # начиная со второй строки (знаем, что первая - шапка таблицы), для каждой строки файла прогоняем:
    l = start_label
    for k in range(2, counter_of_lines):

        # задаём адреса ячеек обрабатываемой строки
        l = col_label + str(k)
        e = col_elapse + str(k)
        ts = col_t_s + str(k)
        er = col_errors + str(k)

        # если имя совпадает с искомым добавляем счётчик, плюсуем сумму
        # проверяем правильность ввода для столбца с временами отклика
        if str(sheet[l].value) == str(name):
            try:
                pass_counter += 1
                summa += int(str(sheet[e].value))
                sum_of_errors += int(str(sheet[er].value))

                # переводим таймстемп в человекочитаемый вид
                timestamp = int(str(sheet[ts].value)) / 1000
                time_format = time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime(timestamp))[11:]

                # добавляем полученные значения в массивы
                arr_of_elapse.append(sheet[e].value)
                arr_of_ts.append(time_format)
            except ValueError:
                print('Ошибка данных в столбце времени отклика или ошибок')
                exit()

    print(name)

    # в массив выводим итоговую информацию по транзакции + вывод в консоль
    main_array = print_info(pass_counter, summa, arr_of_elapse, sum_of_errors)

    #
    for k in range(1, 77):
        if name == names_sheet['A' + str(k)].value:
            result_sheet['A' + str(counter_for_writing)] = names_sheet['B' + str(k)].value
            result_sheet['B' + str(counter_for_writing)] = names_sheet['C' + str(k)].value

    #
    result_sheet['C' + str(counter_for_writing)] = name
    result_sheet['D' + str(counter_for_writing)] = main_array[0][7:]
    result_sheet['E' + str(counter_for_writing)] = main_array[1][9:]
    result_sheet['F' + str(counter_for_writing)] = main_array[4][5:]
    result_sheet['G' + str(counter_for_writing)] = main_array[5][5:]
    result_sheet['H' + str(counter_for_writing)] = main_array[2][14:]
    result_sheet['I' + str(counter_for_writing)] = main_array[3][14:]
    result_sheet['J' + str(counter_for_writing)] = main_array[6][8:]

    counter_for_writing += 1

    # настраиваем вид графика, данные и тд
    plt.style.use('seaborn-colorblind')

    plt.figure(figsize=(12, 6), facecolor='aliceblue')

    # для графика точками plt.scatter
    plt.plot(arr_of_ts, arr_of_elapse, label='Время отклика')

    plt.suptitle(name)
    plt.subplots_adjust(left=0.06, bottom=0.15, top=0.94, right=0.8)  # , right=0.8
    plt.xticks(arr_of_ts[::frequency], rotation=90)
    plt.grid(axis='both', alpha=.2)
    plt.legend(loc='upper left')
    plt.figtext(0.83, 0.8, main_array[0])
    plt.figtext(0.83, 0.7, main_array[1])
    plt.figtext(0.83, 0.6, main_array[2])
    plt.figtext(0.83, 0.5, main_array[3])
    plt.figtext(0.83, 0.4, main_array[4])
    plt.figtext(0.83, 0.3, main_array[5])
    plt.figtext(0.83, 0.2, main_array[6])

    # вывод графика в окно
    # plt.show()

    # сохранение графика в папку проекта и закрытие графика
    way = del_slashes(del_question(name[0:30])) + '_' + str(name_of_sheet) + '_' + str(date) + '.png'
    print(way)
    plt.savefig(way, bbox_inches='tight')
    plt.close()

# закрываем и удаляем временный файл
my_file.close()
os.remove(temp_file)

#
result_file.save('result_file_' + str(name_of_sheet) + '_' + str(date) + '.xlsx')
